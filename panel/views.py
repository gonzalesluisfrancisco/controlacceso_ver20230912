from django.core import serializers
from django.http.request import HttpRequest
from django.http.response import HttpResponse
from django.shortcuts import render, redirect
from django.http import JsonResponse

from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth import get_user_model
from django.utils.encoding import force_str

from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.contrib.sites.shortcuts import get_current_site
from django.utils.encoding import force_bytes

from . import models

# from datetime import datetime
from django.utils import timezone
from django.utils.timezone import activate
import pytz

from django.db.models import Q

from . import Funciones as Fun
from .Funciones import debugPrint
from django.core.exceptions import ObjectDoesNotExist

import re
import random
import string
from datetime import datetime
import traceback
import openpyxl

from random import randrange
import json

# Create your views here.
TEMPLATE_DIRS = (
    'os.path.join(BASE_DIR, "templates")'
)

@login_required(login_url='autenticacion')
def dashboard(request):
    debugPrint("Ingreso a dashboard")
    X_var = 'Personal'
    Y_var = 'Ingreso/Salidas'
    #Color_var = {role: 'style'}
    #values = [[X_var, Y_var]]
    #debugPrint(models.LiveData.objects.count())
    values=[["PERSONAL QUE INGRESO",models.LiveData.objects.count(),'red']]
    values.append(["PERSONAL QUE SALIO",randrange(100),'blue'])
    values_JSON = json.dumps(values)
    return render(request, "dashboard/dashboard.html", {'values':values_JSON})

def home_view(request: HttpRequest) -> HttpResponse:
    return render(request, 'home.html')


def index_noautenticado(request):

    return render(request, "index.html")


def autenticacion(request):
    if request.method == "POST":

        if request.POST.get("Comando") == "VerificarLogin":
            Usuario = request.POST.get("Usuario")
            Contrasena = request.POST.get("Contrasena")
            Captcha = request.POST.get("Captcha")
            FormatoValido, Mensaje = Fun.FormatoLoginValidos(
                Usuario, Contrasena, Captcha)
            if not FormatoValido:
                return JsonResponse({"Estado": "Invalido", 'Mensaje': Mensaje})
            DatosValidos, Mensaje = Fun.DatosLoginValidos(
                request, Usuario, Contrasena)
            if DatosValidos:
                Correo = Fun.EnviarToken(Usuario, Contrasena, request)
                return JsonResponse({"Estado": "Valido", 'Mensaje': Mensaje, "Correo": Correo})
            return JsonResponse({"Estado": "Invalido", 'Mensaje': Mensaje})

        elif request.POST.get("Comando") == "VerificarToken":
            Usuario = request.POST.get("Usuario")
            Contrasena = request.POST.get("Contrasena")
            Token = request.POST.get("Token")
            if len(Token) != 6:
                return JsonResponse({"Estado": "Invalido", "Mensaje": "El Token debe ser de 8 digitos"})
            if not Fun.VerificarToken(Token):
                return JsonResponse({"Estado": "Invalido", 'Mensaje': "El token ingresado no es correcto"})
            DatosValidos, Mensaje = Fun.DatosLoginValidos(
                request, Usuario, Contrasena)
            if not DatosValidos:
                return JsonResponse({"Estado": "Invalido", 'Mensaje': "No deberias poder ver esto"})
            login(request, authenticate(
                request, username=Usuario, password=Contrasena))
            return JsonResponse({"Estado": "Valido", "Mensaje": "El token ha sido validado correctamente"})

        elif request.POST.get("Comando") == "RecuperarCuenta":
            Correo = request.POST.get("Correo")
            if len(Correo) == 0:
                return JsonResponse({"Estado": "Invalido", "Mensaje": "Debe ingresar su correo electronico"})
            try:
                Usuario = User.objects.get(email=Correo)
            except ObjectDoesNotExist:
                return JsonResponse({"Estado": "Invalido", "Mensaje": "El correo ingresado no se encuentra registrado"})
            token = default_token_generator.make_token(Usuario)
            uid = urlsafe_base64_encode(force_bytes(Usuario.pk))
            current_site = get_current_site(request)
            Asunto = 'Recuperar contraseña - DIACSA'
            reset_url = f"https://{current_site}/reset-password/{uid}/{token}/"
            MensajeHTML = f"""\
            <html>
            <head></head>
            <body>
                <p>Hola, <span style="font-size: larger;">{Usuario.first_name}</span>!</p>
                <p>Entra a este enlace para recuperar tu contraseña: <br><span style="font-size: larger;"><b>{reset_url}</b></span></p>
            </body>
            </html>
            """
            Fun.EnviaCorreo(Usuario.email, Asunto, MensajeHTML)
            return JsonResponse({"Estado": "Valido", "Mensaje": "El enlace de recuperacion ha sido enviado a su correo"})

        else:
            return JsonResponse({"Estado": "Invalido", 'Mensaje': "El comando es desconocido"})

    signout(request)
    return render(request, "login.html")


def reset_password(request, uidb64, token):
    if request.method == 'POST':
        try:
            UID = force_str(urlsafe_base64_decode(uidb64))
            Usuario = User.objects.get(pk=UID)
        except (TypeError, ValueError, OverflowError, User.DoesNotExist):
            Usuario = None
        if Usuario is None or not default_token_generator.check_token(Usuario, token):
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El enlace de recuperacion no es valido"})
        NuevaContrasena = request.POST.get('Contrasena1')
        ConfirmacionContrasena = request.POST.get('Contrasena2')
        if NuevaContrasena != ConfirmacionContrasena:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "Las contraseñas ingresadas no coinciden"})
        if not Fun.ContrasenaEsFuerte(NuevaContrasena):
            return JsonResponse({"Estado": "Invalido", "Mensaje": "La nueva contraseña debe tener:\nUna mayuscula.\nUna minuscula.\nSin numeros consecutivos.\nMinimo 8 caracteres"})
        Usuario.set_password(NuevaContrasena)
        Usuario.save()
        return JsonResponse({"Estado": "Valido", "Mensaje": "La contraseña se ha cambiado correctamente"})
    else:
        return render(request, 'reset_password.html')


@login_required(login_url='autenticacion')
def signout(request):
    logout(request)
    return redirect("autenticacion")


@login_required(login_url='autenticacion')
def index(request):
    debugPrint("Ingreso a intentar leer el livedata")
    users = models.LiveData.objects.all()
    data = {'total': users.count()}
    return render(request, "index.html", data)


@login_required(login_url='autenticacion')
def listar(request):
    if request.method == "POST" and request.user.is_superuser and (request.POST.get("Comando") == "Agregar"):
        debugPrint("Entra a Agregar Plantilla por excel")
        header = ["N° PERSONA", "APELLIDO PATERNO", "APELLIDO MATERNO", "NOMBRE", "DNI", "FECHA DE NACIMIENTO", "PROYECTO", "CENTRO DE COSTE", "TIPO DE TRABAJADOR",
                  "CLAVE DE SEXO", "FECHA DE ALTA", "FECHA DE BAJA", "MOTIVO DE CESE", "CARGO", "CÓDIGO DE TARJETA", "ÁREA", "SERVICIO", "SUPERVISIÓN", "GUARDIA", "CORREO", "N° CELULAR"]
        n_columnas = len(header)
        ExcelFile = request.FILES.get("excelFile")
        if not ExcelFile:
            return JsonResponse({'Estado': 'Invalido', 'Mensaje': "No se ha podido leer el archivo"})
        try:
            Excel = openpyxl.load_workbook(
                ExcelFile, read_only=True, keep_vba=True, data_only=True, keep_links=False)
        except Exception as e:
            return JsonResponse({'Estado': 'Invalido', 'Mensaje': str(e) + "\n" + traceback.format_exc()})
        Hoja = Excel.active
        if not Hoja:
            return JsonResponse({'Estado': 'Invalido', 'Mensaje': "El archivo no tienen ninguna hoja"})
        FilasTotales = Hoja.max_row
        debugPrint("La integridad del archivo es correcta")
        for i in range(0, n_columnas):
            try:
                if not str(Hoja[1][i].value).strip() == header[i].strip():
                    debugPrint(
                        f"{Hoja[1][i].value} != {header[i]}")
                    return JsonResponse({'Mensaje': 'La cabecera es incorrecta (parcialmente)', 'Estado': "Invalido"})
            except:
                return JsonResponse({'Mensaje': 'La cabecera es incorrecta', 'Estado': "Invalido"})
            debugPrint(
                f"Cabecera {i}, correcta: {Hoja[1][i].value}")
        debugPrint("La cabecera es correcta")
        filasInvalidas = [0]
        for i in range(2, FilasTotales):
            Fila = Hoja[i]
            for i in range(0, n_columnas):
                debugPrint(f"{header[i]}: {Fila[i].value}")
            return JsonResponse({'Mensaje': 'Planilla actualizada correctamente', 'Estado': "Valido"})
        if request.method == "POST" and request.user.is_superuser and (request.POST.get("Comando") == "Agregar"):
            ExcelFile = request.FILES.get("excelFile")
            if ExcelFile:
                try:
                    Excel = openpyxl.load_workbook(
                        ExcelFile, read_only=True, keep_vba=True, data_only=True, keep_links=False)
                    Hoja = Excel.active
                    FilasTotales = Hoja.max_row
                    debugPrint("Llego a leer el excel, eliminar")
                    for i in range(0, n_columnas):
                        try:
                            if not str(Hoja[1][i].value).strip() == header[i].strip():
                                debugPrint(
                                    f"{Hoja[1][i].value} != {header[i]}")
                                return JsonResponse({'Mensaje': 'La cabecera es incorrecta (parcialmente)', 'Estado': "Invalido"})
                        except:
                            return JsonResponse({'Mensaje': 'La cabecera es incorrecta', 'Estado': "Invalido"})
                        debugPrint(
                            f"Cabecera {i}, correcta: {Hoja[1][i].value}")
                    debugPrint("Hasta aqui cabecera correcta")
                    for i in range(2, FilasTotales):
                        None
                except Exception as e:
                    return JsonResponse({'Estado': 'Invalido', 'Mensaje': str(e) + "\n" + traceback.format_exc()})
        debugPrint("El archivo no se ha podido leer")
        return JsonResponse({"Estado": "Invalido", "Mensaje": "El archivo no se ha podido leer"})

    if request.method == "POST" and request.POST.get("Comando") == "eliminarPlantilla":
        if not request.user.is_superuser and not request.user.is_staff:
            return JsonResponse({"Mensaje": "Su cuenta no tiene los privilegios para realizar esta accion"})
        tablaPlantilla = models.PersonalRegistrado.objects.all()
        tablaPlantilla.delete()
        return JsonResponse({"Mensaje": "La plantilla ha sido borrada correctamente", "Estado": "Valido"})
    users = models.PersonalRegistrado.objects.all()
    datos = {'personalregistrado': users}
    if request.user.is_staff:
        return render(request, "crud_aesadiacsa/listar.html", datos)
    return render(request, "crud_aesadiacsa/listar_basic.html", datos)


@login_required(login_url='autenticacion')
def agregar(request):
    if request.method == 'POST':
        debugPrint(request.POST.get('cardid'))
        debugPrint(request.POST.get('nombre'))
        debugPrint(request.POST.get('apellido'))
        debugPrint(request.POST.get('empresa'))
        debugPrint(request.POST.get('cargo'))
        debugPrint(request.POST.get('telefono'))
        debugPrint(request.POST.get('correo'))
        debugPrint(request.POST.get('f_nac'))
        # agregar datos
        if request.POST.get('cardid') and request.POST.get('nombre') and request.POST.get('apellido') and request.POST.get('telefono') and request.POST.get('correo') and request.POST.get('f_nac'):
            users = models.PersonalRegistrado.objects.all()
            cantidadactualRegistrada = users.count()
            user = models.PersonalRegistrado()
            user.id = cantidadactualRegistrada+1
            user.cardidHex = request.POST.get('cardid')
            user.nombre = request.POST.get('nombre')
            user.apellido = request.POST.get('apellido')
            user.empresa = request.POST.get('empresa')
            user.cargo = request.POST.get('cargo')
            user.correo = request.POST.get('correo')
            user.telefono = request.POST.get('telefono')
            user.f_nac = request.POST.get('f_nac')
            user.f_registro = timezone.now()
            user.save()
            return redirect('listar')
        datos = {'r2': "Debe ingresar todos los campos correctamente"}
        return render(request, "crud_aesadiacsa/agregar.html", datos)

    else:
        return render(request, "crud_aesadiacsa/agregar.html")


@login_required(login_url='autenticacion')
def actualizar(request, codigo):
    if request.method == 'POST':
        # debugPrint(request.POST.get('id'))
        debugPrint(request.POST.get('cardid'))
        debugPrint(request.POST.get('nombre'))
        debugPrint(request.POST.get('apellido'))
        debugPrint(request.POST.get('empresa'))
        debugPrint(request.POST.get('cargo'))
        debugPrint(request.POST.get('telefono'))
        debugPrint(request.POST.get('correo'))
        debugPrint(request.POST.get('f_nac'))
        # agregar datos
        if request.POST.get('cardid') and request.POST.get('id') and request.POST.get('nombre') and request.POST.get('apellido') and request.POST.get('telefono') and request.POST.get('correo') and request.POST.get('f_nac'):
            user = models.PersonalRegistrado()
            # user.id = request.POST.get('id')
            user.cardidHex = request.POST.get('cardid')
            user.nombre = request.POST.get('nombre')
            user.apellido = request.POST.get('apellido')
            user.empresa = request.POST.get('empresa')
            user.cargo = request.POST.get('cargo')
            user.correo = request.POST.get('correo')
            user.telefono = request.POST.get('telefono')
            user.f_nac = request.POST.get('f_nac')
            user.f_registro = timezone.now()
            user.save()
            return redirect('listar')
        datos = {'r2': "Debe ingresar todos los campos correctamente"}
        return render(request, "crud_aesadiacsa/actualizar.html", datos)
    else:
        datosuser = models.PersonalRegistrado.objects.get(cardidHex=codigo)
        debugPrint("Obtuvo datos de usuario")
        debugPrint(datosuser)
        datos = {'personalregistrado': datosuser}

        return render(request, "crud_aesadiacsa/actualizar.html", datos)


@login_required(login_url='autenticacion')
def eliminar(request, codigo):
    tupla = models.PersonalRegistrado.objects.get(cardidHex=codigo)
    tupla.delete()
    return redirect('listar')


@login_required(login_url='autenticacion')
def livedata(request):
    if request.method == "POST" and request.POST.get("Comando") == "TablaLiveData":
        debugPrint("Tabla LiveData")
        debugPrint("Search")
        draw = int(request.POST.get('draw', 0))
        start = int(request.POST.get('start', 0))
        length = int(request.POST.get('length', 0))
        search_value = request.POST.get('search[value]', '')
        queryset = models.LiveData.objects.order_by(
            "-f_ingreso", "-h_ingreso", "-id")
        if search_value != '':
            queryset = queryset.filter(
                Q(id__icontains=search_value) |
                Q(ubicacion__icontains=search_value) |
                Q(cardidHex__icontains=search_value) |
                Q(nombre__icontains=search_value) |
                Q(apellido__icontains=search_value) |
                Q(empresa__icontains=search_value) |
                Q(cargo__icontains=search_value) |
                Q(f_ingreso__icontains=search_value) |
                Q(h_ingreso__icontains=search_value)
            ).order_by("-f_ingreso", "-h_ingreso")
        total_records = queryset.count()
        queryset = queryset[start:start+length]
        data = []
        for i, obj in enumerate(queryset, start=0):
            item = {
                'id': str(total_records - start - i),
                'ubicacion': obj.ubicacion,
                'cardidHex': obj.cardidHex,
                'nombre': obj.nombre,
                'apellido': obj.apellido,
                'empresa': obj.empresa,
                'cargo': obj.cargo,
                'f_evento': obj.f_ingreso,
                'h_evento': obj.h_ingreso,
            }
            data.append(item)
        response = {
            "draw": draw,
            "recordsTotal": total_records,
            "recordsFiltered": total_records,
            "data": data,
        }
        # debugPrint(response)
        return JsonResponse(response)
    elif request.method == "POST" and request.POST.get("Comando") == "ObtenerHoraTotal":
        debugPrint("ObtenerHoraTotal")
        lima_timezone = pytz.timezone('America/Lima')
        lima_time = timezone.now().astimezone(
            lima_timezone).strftime('%Y-%m-%d %H:%M:%S')
        total = models.LiveData.objects.order_by(
            "-f_ingreso", "-h_ingreso").count()
        return JsonResponse({"Estado": "Valido", "Total": total, "Hora": lima_time})
    # users = models.LiveData.objects.all()
    # activate(pytz.timezone('America/Lima'))
    # debugPrint(timezone.now())
    # datos = { 'livedata' : users,             'fecha_y_hora': timezone.now(),             'total': users.count()}
    return render(request, "livedata/livedata.html")


@login_required(login_url='autenticacion')
def livedata_llenar(request):
    users = models.LiveData.objects.all()
    # users.
    # datosuser = models.PersonalRegistrado.objects.get(id=codigo)
    return redirect('livedata')


@login_required(login_url='autenticacion')
def livedata_agregar(request):
    debugPrint(request.POST.get('comando'))
    if request.method == 'POST' and request.POST.get('comando') == 'consultaDatos':
        debugPrint(request.POST.get('cardid'))
        debugPrint(request.POST.get('nombre'))
        debugPrint(request.POST.get('apellido'))
        debugPrint(request.POST.get('cargo'))
        debugPrint(request.POST.get('f_ingreso'))
        debugPrint(request.POST.get('h_ingreso'))
        # agregar datos
        if len(request.POST.get('cardid', ' ')) != 8:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El Card ID debe ser de 8 caracteres"})
        elif not request.POST.get('nombre') and not request.POST.get('apellido') and not request.POST.get('cargo') and not request.POST.get('f_ingreso') and not request.POST.get('h_ingreso'):
            try:
                users = models.PersonalRegistrado.objects.get(
                    cardidHex=request.POST.get('cardid'))
            except ObjectDoesNotExist:
                return JsonResponse({"ubicacion": '',
                                    "nombre": '',
                                     "apellido": '',
                                     "cargo": '',
                                     "f_ingreso": '',
                                     "h_ingreso": ''})
            else:
                return JsonResponse({"ubicacion": '',
                                    "nombre": users.nombre,
                                     "apellido": users.apellido,
                                     "cargo": users.cargo,
                                     "f_ingreso": '',
                                     "h_ingreso": ''})

    if request.method == 'POST' and request.POST.get('comando') == 'agregarLivedata':
        debugPrint("Ingreso a POST agreegar livedata")
        ubicacion = request.POST.get('ubicacion')
        carid = request.POST.get('cardid')
        nombre = request.POST.get('nombre')
        apellido = request.POST.get('apellido')
        cargo = request.POST.get('cargo')
        fecha = request.POST.get('f_ingreso')
        hora = request.POST.get('h_ingreso')
        if request.POST.get('cardid') and request.POST.get('nombre') and request.POST.get('apellido') and request.POST.get('cargo') and request.POST.get('f_ingreso') and request.POST.get('h_ingreso'):
            try:
                users = models.LiveData.objects.get(
                    cardidHex=request.POST.get('cardid'))
                datos = {'Estado': "Invalido",
                         'Mensaje': "El Card ID ingresado ya se encuentra al interior de la mina"}
                return JsonResponse(datos)
            except ObjectDoesNotExist:
                None
            cantidadactualRegistrada = models.LiveData.objects.all().count()
            user = models.LiveData()
            user.id = cantidadactualRegistrada+1
            user.cardidHex = carid
            user.nombre = nombre
            user.apellido = apellido
            user.cargo = cargo
            user.f_ingreso = fecha
            user.h_ingreso = hora
            user.save()
            debugPrint("GuardaLivedata")
            user2 = models.Historial()
            cantidadactualRegistrada = models.Historial.objects.all()
            debugPrint("Total: ", cantidadactualRegistrada)
            user2.id = cantidadactualRegistrada + 1
            user2.cardidHex = carid
            user2.nombre = nombre
            user2.apellido = apellido
            user2.cargo = cargo
            user2.f_evento = fecha
            user2.h_evento = hora
            user2.save()
            debugPrint("GuardaHistorial")
            try:
                users = models.PersonalRegistrado.objects.get(
                    nombre=request.POST.get('nombre'), apellido=request.POST.get('apellido'))
            except ObjectDoesNotExist:
                user3 = models.NoRegistrados()
                user3.ubicacion = ubicacion
                user3.cardidHex = carid
                user3.f_evento = fecha
                user3.h_evento = hora
                user3.evento = 'Ingreso'
                user3.save()
                debugPrint("Guarda no registrados")
            return JsonResponse({'Mensaje': "El personal ha sido agregado con exito"})
        datos = {"Estado": "Invalido",
                 'Mensaje': "Debe ingresar todos los campos correctamente"}
        return JsonResponse(datos)

    elif request.method == 'GET':
        return render(request, "livedata/livedata_agregar.html")


@login_required(login_url='autenticacion')
def livedata_eliminar(request):
    if request.method == "POST" and request.POST.get("Comando") == "ObtenerDatosCardID":
        debugPrint(request.POST.get('cardidHex'))
        if request.POST.get('cardidHex'):
            cardidHex_a_borrar = request.POST.get('cardidHex')
            tupla = models.LiveData.objects.get(cardidHex=cardidHex_a_borrar)
            return JsonResponse({"Estado": "Valido",
                                 "ubicacion": tupla.ubicacion,
                                 "cardid": tupla.cardidHex,
                                 "nombre": tupla.nombre,
                                 "apellido": tupla.apellido,
                                 "cargo": tupla.cargo,
                                 "fechaingreso": tupla.f_ingreso,
                                 "horaingreso": tupla.h_ingreso
                                 })
        else:
            return JsonResponse({"Estado": "Invalido"})
    elif request.method == "POST" and request.POST.get("Comando") == "EliminarEntrada":
        if request.POST.get('cardidHex'):
            try:
                cardidHex_a_borrar = request.POST.get('cardidHex')
                tupla = models.LiveData.objects.get(
                    cardidHex=cardidHex_a_borrar)
                tupla.delete()
                return JsonResponse({"Estado": "Valido", "Mensaje": "El registro ha sido eliminado"})
            except Exception as ex:
                return JsonResponse({"Estado": "Invalido", "Mensaje": f"{str(ex)}"})
        return JsonResponse({"Estado": "Invalido", "Mensaje": "El CardID seleccionado no se encuentra en la base de datos"})
    if request.method == "GET":
        users = models.LiveData.objects.all()
        datos = {'livedata': users}
        return render(request, "livedata/livedata_eliminar.html", datos)


@login_required(login_url='autenticacion')
def marcacion(request):
    if request.method == "POST" and request.POST.get("Comando") == "TablaMarcacion":
        debugPrint("Search")
        min = request.POST.get("min")
        max = request.POST.get("max")
        draw = int(request.POST.get('draw', 0))
        start = int(request.POST.get('start', 0))
        length = int(request.POST.get('length', 0))
        search_value = request.POST.get('search[value]', '')
        queryset = models.Historial.objects.order_by(
            '-f_evento', '-h_evento', 'cardidHex')
        if min != "" and max != "":
            queryset = queryset.filter(
                f_evento__range=(min, max)).order_by('-id')
        if search_value != '':
            queryset = queryset.filter(
                Q(id__icontains=search_value) |
                Q(ubicacion__icontains=search_value) |
                Q(cardidHex__icontains=search_value) |
                Q(nombre__icontains=search_value) |
                Q(apellido__icontains=search_value) |
                Q(empresa__icontains=search_value) |
                Q(cargo__icontains=search_value) |
                Q(f_evento__icontains=search_value) |
                Q(h_evento__icontains=search_value) |
                Q(evento__icontains=search_value) |
                Q(status__icontains=search_value)
            ).order_by("-id")
        total_records = queryset.count()
        queryset = queryset[start:start+length]
        data = []
        id = total_records - start + 1
        for obj in queryset:
            id -= 1
            item = {
                'primarykey': obj.id,
                'id': id,
                'ubicacion': obj.ubicacion,
                'cardidHex': obj.cardidHex,
                'nombre': obj.nombre,
                'apellido': obj.apellido,
                'empresa': obj.empresa,
                'cargo': obj.cargo,
                'f_evento': obj.f_evento,
                'h_evento': obj.h_evento,
                'evento': obj.evento,
                'status': obj.status,
            }
            data.append(item)
        response = {
            "draw": draw,
            "recordsTotal": total_records,
            "recordsFiltered": total_records,
            "data": data
        }
        # debugPrint(response)
        return JsonResponse(response)
    elif request.method == "GET" and request.GET.get("Comando") == "DescargarExcel":
        min = request.GET.get('FechaInicial')
        max = request.GET.get('FechaFinal')
        search_value = request.GET.get('Search', '')
        debugPrint(min, max)
        queryset = models.Historial.objects.order_by('-id')
        if min != "" and max != "":
            queryset = queryset.filter(
                f_evento__range=(min, max)).order_by('-id')
        if search_value != '':
            queryset = queryset.filter(
                Q(id__icontains=search_value) |
                Q(ubicacion__icontains=search_value) |
                Q(cardidHex__icontains=search_value) |
                Q(nombre__icontains=search_value) |
                Q(apellido__icontains=search_value) |
                Q(empresa__icontains=search_value) |
                Q(cargo__icontains=search_value) |
                Q(f_evento__icontains=search_value) |
                Q(h_evento__icontains=search_value) |
                Q(evento__icontains=search_value) |
                Q(status__icontains=search_value)
            ).order_by("-id")
        total_records = queryset.count() + 1
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['#', 'Ubicacion', 'Card ID', 'Nombre', 'Apellido',
                  'Cargo', 'F. Evento', 'H. Evento', 'Evento'])
        for item in queryset:
            total_records -= 1
            ws.append([total_records, item.ubicacion, item.cardidHex, item.nombre,
                      item.apellido, item.cargo, item.f_evento, item.h_evento, item.evento, item.status])
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Historial_ingresos_salidas.xlsx"'
        wb.save(response)
        return response
    elif request.method == "POST" and request.POST.get("comando") == "eliminarRegistro":
        if not (request.user.is_staff or request.user.is_superuser):
            return JsonResponse({"Estado": "noExitoso", "Mensaje": "Su cuenta no tiene los permisos para hacer esta accion"})
        debugPrint("ingreso a eliminar registro")
        primarykey = request.POST.get("primarykey")
        registroHistorial = models.Historial.objects.get(id=primarykey)
        if registroHistorial.status == '0':
            registroHistorial.status = "10"
        elif registroHistorial.status == '1':
            registroHistorial.status = "11"
        else:
            debugPrint("respuesta estado, mensaje no exitoso")
            return JsonResponse({"Estado": "noExitoso", "Mensaje": "El registro se encontraba eliminado previamente"})
        try:
            registroLivedata = models.LiveData.objects.get(
                cardidHex=registroHistorial.cardidHex)
            debugPrint("Se encontro en registro en LiveData")
            if registroHistorial.evento == "Ingreso" and registroHistorial.f_evento == registroLivedata.f_ingreso and registroHistorial.h_evento == registroLivedata.h_ingreso:
                ultimoHistorial = models.Historial.objects.filter(
                    Q(cardidHex=registroHistorial.cardidHex) & (Q(status='0') | Q(status='1'))).order_by(
                    '-f_evento', '-h_evento')
                debugPrint("El registro encontrado es ingreso")
                if ultimoHistorial[0] == registroHistorial:
                    try:
                        if ultimoHistorial[1].evento == "Salida":
                            registroLivedata.delete()
                    except:
                        registroLivedata.delete()
        except:
            debugPrint("Except al encontrar el livedata del registro eliminado")
        if registroHistorial.evento == "Salida":
            debugPrint("El registro agregado es salida")
            ultimoHistorial = models.Historial.objects.filter(
                Q(cardidHex=registroHistorial.cardidHex) & (Q(status='0') | Q(status='1'))).order_by(
                '-f_evento', '-h_evento')
            if ultimoHistorial[0] == registroHistorial:
                try:
                    if ultimoHistorial[1].evento == "Ingreso":
                        nuevoLivedata = models.LiveData()
                        nuevoLivedata.cardidHex = registroHistorial.cardidHex
                        nuevoLivedata.nombre = registroHistorial.nombre
                        nuevoLivedata.apellido = registroHistorial.apellido
                        nuevoLivedata.cargo = registroHistorial.cargo
                        nuevoLivedata.empresa = registroHistorial.empresa
                        nuevoLivedata.f_ingreso = ultimoHistorial[1].f_evento
                        nuevoLivedata.h_ingreso = ultimoHistorial[1].h_evento
                        nuevoLivedata.ubicacion = nuevoLivedata.ubicacion
                        nuevoLivedata.save()
                except:
                    nuevoLivedata = models.LiveData()
                    nuevoLivedata.cardidHex = registroHistorial.cardidHex
                    nuevoLivedata.nombre = registroHistorial.nombre
                    nuevoLivedata.apellido = registroHistorial.apellido
                    nuevoLivedata.cargo = registroHistorial.cargo
                    nuevoLivedata.empresa = registroHistorial.empresa
                    nuevoLivedata.f_ingreso = ultimoHistorial[1].f_evento
                    nuevoLivedata.h_ingreso = ultimoHistorial[1].h_evento
                    nuevoLivedata.ubicacion = nuevoLivedata.ubicacion
                    nuevoLivedata.save()
                debugPrint("eliminado con registro agregado salida")
        debugPrint("responde eliminacion correcta")
        registroHistorial.save()
        return JsonResponse({"Mensaje": "El registro se ha eliminado correctamente"})
    debugPrint("Marcacion")
    # users = models.Historial.objects.all().order_by('-id')
    # datos = { 'marcacion' : users}
    # debugPrint(users[0].id, users[0].f_evento, type(users))
    return render(request, "marcacion/marcacion.html")


@login_required(login_url='autenticacion')
def marcacion_agregar(request):
    if not request.user.is_superuser and not request.user.is_staff:
        return render(request, "marcacion/marcacion_agregar_denegado.html")
    debugPrint(request.POST.get('comando'))
    if request.method == 'POST' and request.POST.get('comando') == 'consultaDatos':
        debugPrint(request.POST.get('cardid'))
        debugPrint(request.POST.get('nombre'))
        debugPrint(request.POST.get('apellido'))
        debugPrint(request.POST.get('cargo'))
        debugPrint(request.POST.get('f_ingreso'))
        debugPrint(request.POST.get('h_ingreso'))
        # agregar datos
        if len(request.POST.get('cardid', ' ')) != 8:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El Card ID debe ser de 8 caracteres"})
        elif not request.POST.get('nombre') and not request.POST.get('apellido') and not request.POST.get('cargo') and not request.POST.get('f_ingreso') and not request.POST.get('h_ingreso'):
            try:
                users = models.PersonalRegistrado.objects.get(
                    cardidHex=request.POST.get('cardid'))
            except ObjectDoesNotExist:
                return JsonResponse({"ubicacion": 'San Rafael',
                                    "nombre": 'No Registrado',
                                     "apellido": 'No Registrado',
                                     "empresa": "No Registrado",
                                     "cargo": 'No Registrado',
                                     "f_ingreso": '',
                                     "h_ingreso": ''})
            else:
                return JsonResponse({"ubicacion": 'San Rafael',
                                    "nombre": users.nombre,
                                     "apellido": users.apellido,
                                     "cargo": users.cargo,
                                     "empresa": users.empresa,
                                     "f_ingreso": '',
                                     "h_ingreso": ''})

    if request.method == 'POST' and request.POST.get('comando') == 'agregarLivedata':
        if not (request.user.is_staff or request.user.is_superuser):
            return JsonResponse({"Estado": "noExitoso", "Mensaje": "Su cuenta no tiene los permisos para hacer esta accion"})
        debugPrint("Ingreso a POST agreegar livedata")
        carid = request.POST.get('cardid')
        ubicacion = request.POST.get('ubicacion')
        nombre = request.POST.get('nombre')
        apellido = request.POST.get('apellido')
        empresa = request.POST.get('empresa')
        cargo = request.POST.get('cargo')
        fecha = request.POST.get('f_ingreso')
        hora = request.POST.get('h_ingreso')
        evento = request.POST.get("evento")
        if carid and ubicacion and nombre and apellido and empresa and cargo and fecha and hora and evento:
            user2 = models.Historial()
            cantidadactualRegistrada = models.Historial.objects.all().count()
            user2.id = cantidadactualRegistrada + 1
            user2.cardidHex = carid
            user2.ubicacion = ubicacion
            user2.nombre = nombre
            user2.apellido = apellido
            user2.empresa = empresa
            user2.cargo = cargo
            user2.f_evento = fecha
            user2.h_evento = hora
            user2.evento = evento
            user2.status = "1"
            user2.save()
            debugPrint("GuardaHistorial")
            ultimoHistorial = models.Historial.objects.filter(
                Q(cardidHex=carid) & (Q(status='0') | Q(status='1'))).order_by('-f_evento', '-h_evento')
            if ultimoHistorial[0] != user2:
                debugPrint(
                    "Registro guardado pero no es el ultimo del historial")
                return JsonResponse({"Mensaje": "El registro se ha agregado correctamente"})
            if user2.evento == "Salida":
                try:
                    userLiveData = models.LiveData.objects.get(cardidHex=carid)
                    userLiveData.delete()
                    debugPrint("Se elimino registro de livedata")
                    return JsonResponse({"Mensaje": "Registro agregado correctamente..."})
                except:
                    debugPrint(
                        "Registro guardado pero debio eliminarse de livedata pero no se pudo")
                    return JsonResponse({"Mensaje": "Registro agregado correctamente..."})
            try:
                userLiveData = models.LiveData.objects.get(cardidHex=carid)
                userLiveData.ubicacion = ubicacion
                userLiveData.cardidHex = carid
                userLiveData.nombre = nombre
                userLiveData.apellido = apellido
                userLiveData.empresa = empresa
                userLiveData.cargo = cargo
                userLiveData.f_ingreso = fecha
                userLiveData.h_ingreso = hora
                userLiveData.save()
            except:
                userLiveData = models.LiveData()
                userLiveData.cardidHex = carid
                userLiveData.ubicacion = ubicacion
                userLiveData.cardidHex = carid
                userLiveData.nombre = nombre
                userLiveData.apellido = apellido
                userLiveData.empresa = empresa
                userLiveData.cargo = cargo
                userLiveData.f_ingreso = fecha
                userLiveData.h_ingreso = hora
                userLiveData.save()
            debugPrint("El registro se agrego y se actualizo livedata")
            return JsonResponse({'Mensaje': "El registro se ha agregado correctamente."})
        datos = {"Estado": "Invalido",
                 'Mensaje': "Debe ingresar todos los campos correctamente"}
        return JsonResponse(datos)

    elif request.method == 'GET':
        return render(request, "marcacion/marcacion_agregar.html")


@login_required(login_url='autenticacion')
def noregistrados(request):
    users = models.NoRegistrados.objects.all()
    datos = {'noregistrados': users}
    return render(request, "noregistrados/noregistrados.html", datos)


@login_required(login_url='autenticacion')
def registrarusuario(request):
    if request.method == "POST":
        if request.POST.get("Comando") != "RegistrarUsuario":
            return JsonResponse({"Estado": "Invalido", "Mensaje": "¿Que haces?"})
        Nombre = request.POST.get("Nombre").strip().upper()
        PrimerApellido = request.POST.get("PrimerApellido").strip().upper()
        SegundoApellido = request.POST.get("SegundoApellido").strip().upper()
        DNI = request.POST.get("DNI").strip()
        Correo = request.POST.get("Correo").strip()
        Telefono = request.POST.get("Telefono").strip()
        is_staff = False
        if request.POST.get("Rol") == "Observador y registro de planilla":
            is_staff = True
        if not (len(Nombre) and len(PrimerApellido) and len(SegundoApellido) and len(DNI) and len(Correo) and len(Telefono)):
            return JsonResponse({"Estado": "Invalido", "Mensaje": "Debe llenar todos los campos antes de continuar"})
        if not re.match(r'^[A-Za-z\s]+$', Nombre):
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El nombre debe contener solo letras"})
        if not re.match(r'^[A-Za-z]+$', PrimerApellido):
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El apellido debe contener solo letras"})
        if not re.match(r'^[A-Za-z\s]+$', SegundoApellido):
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El apellido debe contener solo letras"})
        if len(DNI) != 8:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El DNI debe tener 8 caracteres"})
        if not re.match(r'^\d+$', DNI):
            return JsonResponse({"Estado": "Invalido", "Mensaje": "Debe ingresar un dni valido"})
        if models.UserInfo.objects.filter(DNI=DNI).exists():
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El DNI ya se encuentra registrado"})
        if not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', Correo):
            return JsonResponse({"Estado": "Invalido", "Mensaje": "Debe ingresar un correo valido"})
        if User.objects.filter(email=Correo).exists():
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El correo ya se encuentra registrado"})
        if len(Telefono) != 9:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El telefono debe tener 9 caracteres"})
        if not re.match(r'^\d+$', Telefono):
            return JsonResponse({"Estado": "Invalido", "Mensaje": "Debe ingresar un telefono valido"})
        try:
            Username = Nombre[0].lower() + PrimerApellido.lower() + DNI[4:8]
            if User.objects.filter(username=Username).exists():
                Username = Nombre[0].lower() + \
                    PrimerApellido.lower() + DNI[0:4]
            random.seed(int(datetime.now().timestamp()))
            Password = ''.join(random.choice(
                string.ascii_letters + string.digits) for _ in range(10))
            NuevoUsuario = User(username=Username, email=Correo,
                                first_name=Nombre, last_name=PrimerApellido, is_staff=is_staff)
            NuevoUsuario.set_password(Password)
            MensajeHTML = f"""\
            <html>
            <head></head>
            <body>
                <p>Hola, <span style="font-size: larger;">{NuevoUsuario.first_name}</span>!</p>
                <p>Se ha registrado su cuenta, sus datos de acceso son:</p>
                <p><br><span style="font-size: larger;"><b>Usuario: {NuevoUsuario.username}</b></span><br>
                <p><br><span style="font-size: larger;"><b>Contraseña: {Password}</b></span></p>
            </body>
            </html>
            """
            Fun.EnviaCorreo(
                NuevoUsuario.email, "Cuenta registrada - Control de acceso", MensajeHTML)
            NuevoUsuario.save()
            NuevoUsuarioInfo = models.UserInfo(
                User=NuevoUsuario, DNI=DNI, Telefono=Telefono, SegundoApellido=SegundoApellido)
            NuevoUsuarioInfo.save()
            debugPrint(Username, Password)
            return JsonResponse({"Estado": "Valido", "Mensaje": "Registrado correctamente, se ha enviado las credenciales al correo electronico"})
        except Exception as e:
            return JsonResponse({"Estado": "Invalido", "Mensaje": f"{str(e)}"})

    else:
        if request.user.is_superuser:
            return render(request, "registrarusuario/plantillaregistro.html")
        return render(request, "registrarusuario/plantilladenegado.html")


def eliminarusuario(request):
    if request.method == "POST":
        if request.POST.get("Comando") == "ConsultarDatos":
            Usuario = request.POST.get("Usuario")
            if not User.objects.filter(username=Usuario).exists():
                return JsonResponse({"Estado": "Invalido"})
            user = User.objects.get(username=Usuario)
            user2 = models.UserInfo.objects.get(User=user)
            Nombre = user.first_name
            PrimerApellido = user.last_name
            SegundoApellido = user2.SegundoApellido
            DNI = user2.DNI
            Correo = user.email
            Telefono = user2.Telefono
            Rol = "Solo observador"
            if user.is_staff:
                Rol = "Observador y registro de planilla"
            Data = {
                "Estado": "Valido",
                "Nombre": Nombre,
                "PrimerApellido": PrimerApellido,
                "SegundoApellido": SegundoApellido,
                "DNI": DNI,
                "Correo": Correo,
                "Telefono": Telefono,
                "Rol": Rol,
            }
            return JsonResponse(Data)
        elif request.POST.get("Comando") == "EliminarUsuario":
            try:
                Usuario = request.POST.get("Usuario")
                user = User.objects.get(username=Usuario)
                user.delete()
                return JsonResponse({"Estado": "Valido", "Mensaje": "El usuario se elimino correctamente"})
            except Exception as e:
                return JsonResponse({"Estado": "Invalido", "Mensaje": f"{str(e)}"})

    if request.user.is_superuser:
        Usuarios = User.objects.exclude(username=request.user.username)
        return render(request, "eliminarusuario/plantillaeliminar.html", {"Usuarios": Usuarios})
    return render(request, "eliminarusuario/plantilladenegado.html")
